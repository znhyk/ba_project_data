import requests
import csv
from openpyxl import load_workbook

def naver_api_jibun_coord(query):
    # 요청 URL
    url = "https://naveropenapi.apigw.ntruss.com/map-geocode/v2/geocode"

    # 필요한 정보 입력
    #query = "서울특별시 동대문구 답십리동 260"
    #coordinate = "127.1054328,37.3595963"
    client_id = "{삭제}"
    client_secret = "{삭제}"

    # 요청 헤더 설정
    headers = {
        "X-NCP-APIGW-API-KEY-ID": client_id,
        "X-NCP-APIGW-API-KEY": client_secret,
        "Accept": "application/json"
    }

    # 요청 파라미터 설정
    params = {
        "query": query,
        #"coordinate": coordinate #검색중심좌표로, 넣으면 해당 위치로부터의 거리가 계산됨.
    }

    # API 요청 보내기
    response = requests.get(url, params=params, headers=headers)

    # 응답 확인
    if response.status_code == 200:
        result = response.json()
        # 결과 출력
        print("Status:", result["status"])
        #print("Total Count:", result["meta"]["totalCount"])
        if result["meta"]["totalCount"] > 0:
            address = result["addresses"][0]
    else:
        print("Error:", response.status_code, response.text)
    return [address["x"],address["y"]]



# 엑셀 파일 경로 (실제 파일 경로에 맞게 수정해주세요)
excel_file_path = "연립다세대(전월세).xlsx"
# 엑셀 파일 로드
workbook = load_workbook(excel_file_path)
# 첫 번째 시트 선택
sheet = workbook.active

# 각 행 반복해서 출력
row_index = 0
for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더로 간주하고 건너뜁니다.
    csvline=[]
    print(f"Row:{row_index}")
    row_index += 1
    jibun = f'{row[0]} {row[1]}'
    try:
        coords = naver_api_jibun_coord(jibun)
    except:
        continue
    x = coords[0]
    y = coords[1]
    csvline = [x,y,jibun]
    for i in range(2,len(row)):
        csvline.append(row[i])
    fw = open('result_dasedae.csv','a',newline='', encoding='utf-8-sig')
    wr = csv.writer(fw)
    wr.writerow(csvline)
    fw.close()
#workbook.save("result_excel_file.xlsx")
